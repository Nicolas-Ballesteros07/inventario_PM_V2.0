import os
import openpyxl
from openpyxl.utils import column_index_from_string
import xlrd
from xlrd import xldate_as_datetime
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from .models import *
from django.core.cache import cache
from .utils_cache import CACHE_KEY_BAJO_DURACION, CACHE_TIMEOUT
from openpyxl import load_workbook
from django.db import transaction
from datetime import datetime, date
from django.db.models import Prefetch

# ─────────────────────────────────────────────
# Parseo de fechas unificado (usado en compras y vencimientos)
# ─────────────────────────────────────────────
def _parsear_fecha(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    valor = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(valor, formato).date()
        except ValueError:
            continue
    return None

#========================================================
# UTILIDADES PARA PROCESAR LOS ARCHIVOS DE EXCEL DEL INFORME
#========================================================
def leer_excel(archivo, fila_encabezado=2):
    """
    archivo: objeto tipo InMemoryUploadedFile / TemporaryUploadedFile
             (el que llega en request.FILES), NO una ruta en disco.
    fila_encabezado: número de fila (1-indexado) donde están los títulos de columna.
    Los datos se leen desde la fila siguiente.
    """
    nombre = archivo.name
    extension = os.path.splitext(nombre)[1].lower()
    archivo.seek(0)
    contenido = archivo.read()

    if extension == '.xlsx':
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
        ws = wb.active
        headers = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=fila_encabezado, column=col).value
            if valor:
                headers[str(valor).strip().upper()] = col - 1
        filas = []
        for row in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
            fila = [celda if celda is not None else '' for celda in row]
            filas.append(fila)
        wb.close()
        return headers, filas

    elif extension == '.xls':
        wb = xlrd.open_workbook(file_contents=contenido)
        ws = wb.sheet_by_index(0)
        headers = {}
        fila_idx_0based = fila_encabezado - 1
        for col in range(ws.ncols):
            valor = ws.cell_value(fila_idx_0based, col)
            if valor:
                headers[str(valor).strip().upper()] = col
        filas = []
        for row_idx in range(fila_idx_0based + 1, ws.nrows):
            fila = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
            filas.append(fila)
        return headers, filas

    else:
        raise ValueError("Formato no soportado. Use archivos .xls o .xlsx")


def limitar_decimal(valor, max_digits, decimal_places):
    """
    Capa un valor para que siempre quepa en un DecimalField con los parámetros dados.
    """
    try:
        limite = Decimal(10) ** (max_digits - decimal_places) - Decimal('0.' + '0'*(decimal_places-1) + '1')
    except InvalidOperation:
        limite = Decimal(10) ** (max_digits - decimal_places)
    if valor > float(limite):
        return float(limite)
    if valor < float(-limite):
        return float(-limite)
    return float(valor)


def calcular_periodo_meses(fecha_inicio, fecha_fin, max_meses=36):
    """
    Determina cuántos meses completos cubre el rango [fecha_inicio, fecha_fin],
    sin exigir que empiece el día 1. Acepta dos convenciones de 'mes completo':

      A) Día-a-día-anterior: desde el día D hasta el día (D-1) del mes
         siguiente. Ej: 01/08 a 31/08 (1 mes), 15/08 a 14/09 (1 mes).
      B) Día-a-mismo-día: desde el día D hasta el mismo día D, N meses
         después (fin inclusive). Ej: 16/04 a 16/06 (2 meses).

    Ambas respetan meses de 28, 29, 30 o 31 días (vía relativedelta).

    Retorna el número de meses si encuentra una coincidencia exacta en
    cualquiera de las dos convenciones. Si no hay coincidencia, levanta
    ValueError indicando, como referencia, cuál sería la fecha de fin
    correcta para el número de meses más cercano al rango recibido.
    """
    for n in range(1, max_meses + 1):
        fin_convencion_a = fecha_inicio + relativedelta(months=n) - timedelta(days=1)
        fin_convencion_b = fecha_inicio + relativedelta(months=n)
        if fecha_fin.date() in (fin_convencion_a.date(), fin_convencion_b.date()):
            return n

    mejor_n, mejor_fin, menor_diferencia = None, None, None
    for n in range(1, max_meses + 1):
        fin_candidato = fecha_inicio + relativedelta(months=n)
        diferencia = abs((fin_candidato.date() - fecha_fin.date()).days)
        if menor_diferencia is None or diferencia < menor_diferencia:
            mejor_n, mejor_fin, menor_diferencia = n, fin_candidato, diferencia

    raise ValueError(
        'El rango de fechas debe cubrir uno o más meses completos '
        '(por ejemplo, del 15/08 al 14/09, del 16/04 al 16/06, o del 1 '
        'al 30/31 de un mes) para poder calcular la rotación '
        'correctamente. El archivo de ventas tiene el rango '
        f'{fecha_inicio.strftime("%d/%m/%Y")} - '
        f'{fecha_fin.strftime("%d/%m/%Y")}. Para que sean {mejor_n} '
        f'mes(es) completo(s) a partir del '
        f'{fecha_inicio.strftime("%d/%m/%Y")}, el archivo debería '
        f'terminar el {mejor_fin.strftime("%d/%m/%Y")} (o el '
        f'{(mejor_fin - timedelta(days=1)).strftime("%d/%m/%Y")}, según '
        'la convención de cierre que use tu reporte).'
    )


def procesar_archivos(archivo_listado, archivo_ventas):
    """
    Procesa los dos archivos subidos y crea (o REEMPLAZA) el registro
    de CargaReporte correspondiente a ese período.

    Si ya existe una carga con el MISMO periodo_inicio y periodo_fin
    (es decir, se subió el mismo archivo de ventas otra vez, o uno con
    el mismo rango de fechas), se borran los DetalleInforme anteriores
    de esa carga y se recrean con los datos nuevos — así se actualiza
    en vez de duplicar.

    Si el período es distinto (aunque se solape parcialmente, como
    12/06-12/07 vs 12/06-12/08), se trata como una carga nueva e
    independiente, ya que representan cálculos de rotación distintos.
    """
    # 2. Leer listado básico primero (no depende de fechas)
    headers_listado, filas_listado = leer_excel(archivo_listado)

    col_codigo = headers_listado['CODIGO MANTIS']
    col_barras = headers_listado.get('CODIGO BARRAS')
    col_descripcion = headers_listado['DESCRIPCION PRODUCTO']
    col_laboratorio = headers_listado['LABORATORIO']
    col_stock = headers_listado['SALDO DE INVENTARIO']
    col_costo_prom = headers_listado['COSTO PROMEDIO']
    col_precio_venta = headers_listado['PRECIO VENTA ANTES IVA']

    # 3. Leer archivo de ventas (encabezados en fila 3, datos desde fila 4)
    headers_ventas, filas_ventas = leer_excel(archivo_ventas, fila_encabezado=3)

    archivo_ventas.seek(0)
    contenido_ventas = archivo_ventas.read()
    extension_ventas = os.path.splitext(archivo_ventas.name)[1].lower()

    if extension_ventas == '.xlsx':
        wb = openpyxl.load_workbook(BytesIO(contenido_ventas), data_only=True)
        ws = wb.active
        fecha_inicio = ws['C2'].value
        fecha_fin = ws['E2'].value
        wb.close()
    else:
        wb = xlrd.open_workbook(file_contents=contenido_ventas)
        ws = wb.sheet_by_index(0)
        raw_inicio = ws.cell_value(1, 2)
        raw_fin = ws.cell_value(1, 4)
        datemode = wb.datemode
        fecha_inicio = xldate_as_datetime(raw_inicio, datemode) if isinstance(raw_inicio, (int, float)) else raw_inicio
        fecha_fin = xldate_as_datetime(raw_fin, datemode) if isinstance(raw_fin, (int, float)) else raw_fin

    if isinstance(fecha_inicio, datetime) and isinstance(fecha_fin, datetime):
        periodo_meses = calcular_periodo_meses(fecha_inicio, fecha_fin)
    else:
        raise ValueError(
            'No se pudieron leer las fechas de inicio y fin del archivo de ventas.'
        )

    fecha_inicio_date = fecha_inicio.date()
    fecha_fin_date = fecha_fin.date()

    carga_existente = (
        CargaReporte.objects
        .filter(detalles__periodo_inicio=fecha_inicio_date, detalles__periodo_fin=fecha_fin_date)
        .distinct()
        .first()
    )

    if carga_existente:
        carga = carga_existente
        carga.detalles.all().delete()
    else:
        carga = CargaReporte.objects.create(tipo_carga='CARGA_INFORME')

    col_venta_codigo = headers_ventas['CODIGO']
    col_venta_unidades = headers_ventas['UND']

    ventas_dict = {}
    for fila in filas_ventas:
        cod = str(fila[col_venta_codigo]).strip() if fila[col_venta_codigo] else ''
        und = float(fila[col_venta_unidades]) if fila[col_venta_unidades] else 0
        if cod:
            ventas_dict[cod] = und

    # --- 1. Parsear todas las filas del listado a memoria, sin tocar la BD ---
    filas_parseadas = []
    codigos_vistos = set()

    for fila in filas_listado:
        codigo = str(fila[col_codigo]).strip() if fila[col_codigo] else ''
        if not codigo:
            continue

        barras = fila[col_barras] if col_barras is not None else ''
        if isinstance(barras, float):
            barras = str(int(barras))
        else:
            barras = str(barras).strip()

        descripcion = fila[col_descripcion] if col_descripcion < len(fila) else ''
        laboratorio = fila[col_laboratorio] if col_laboratorio < len(fila) else ''

        filas_parseadas.append({
            'codigo': codigo,
            'barras': barras,
            'descripcion': descripcion,
            'laboratorio': laboratorio,
            'stock': float(fila[col_stock]) if fila[col_stock] else 0,
            'costo_prom': float(fila[col_costo_prom]) if fila[col_costo_prom] else 0,
            'precio_venta': float(fila[col_precio_venta]) if fila[col_precio_venta] else 0,
        })
        codigos_vistos.add(codigo)

    if not filas_parseadas:
        return carga

    # --- 2. Traer productos existentes en una sola consulta ---
    productos_existentes = {
        p.codigo_mantis: p
        for p in Producto.objects.filter(codigo_mantis__in=codigos_vistos)
    }

    productos_nuevos = {}
    productos_a_actualizar = []

    for f in filas_parseadas:
        codigo = f['codigo']
        producto = productos_existentes.get(codigo)

        if producto is None:
            if codigo not in productos_nuevos:
                productos_nuevos[codigo] = Producto(
                    codigo_mantis=codigo,
                    barras=f['barras'],
                    descripcion=f['descripcion'],
                    laboratorio=f['laboratorio'],
                    estado=True,
                )
        else:
            producto.barras = f['barras']
            producto.descripcion = f['descripcion']
            producto.laboratorio = f['laboratorio']
            productos_a_actualizar.append(producto)

    # --- 3. Crear/actualizar productos en bloque ---
    with transaction.atomic():
        if productos_nuevos:
            # Producto usa UUID como PK con default en el propio objeto,
            # así que ya tienen su id_random asignado antes de insertar.
            Producto.objects.bulk_create(productos_nuevos.values(), batch_size=500)
            productos_existentes.update(productos_nuevos)
        if productos_a_actualizar:
            Producto.objects.bulk_update(
                productos_a_actualizar,
                ['barras', 'descripcion', 'laboratorio'],
                batch_size=500,
            )

    # --- 4. Armar los DetalleInforme en memoria y guardarlos en bloque ---
    detalles_nuevos = []

    for f in filas_parseadas:
        codigo = f['codigo']
        producto = productos_existentes[codigo]

        stock = f['stock']
        costo_prom = f['costo_prom']
        precio_venta = f['precio_venta']

        rentabilidad = ((precio_venta - costo_prom) / precio_venta * 100) if precio_venta > 0.01 else 0
        rentabilidad = limitar_decimal(rentabilidad, 8, 4)

        unidades_vendidas = ventas_dict.get(codigo, 0)
        rotacion = unidades_vendidas / periodo_meses if periodo_meses > 0 else 0
        rotacion = limitar_decimal(rotacion, 12, 4)

        costo = costo_prom
        duracion = stock / rotacion if rotacion != 0 else 0
        duracion = limitar_decimal(duracion, 12, 4)
        solicitar = limitar_decimal(rotacion - stock, 12, 2)

        if rotacion == 0:
            ind_duracion = "SIN ROTACIÓN"
        elif stock <= 0:
            ind_duracion = "SIN STOCK"
        elif duracion <= 0.5:
            ind_duracion = "BAJO STOCK"
        elif duracion < 2:
            ind_duracion = "OK"
        else:
            ind_duracion = "SOBRE STOCK"

        detalles_nuevos.append(DetalleInforme(
            producto=producto,
            carga=carga,
            stock=stock,
            rentabilidad=rentabilidad,
            rotacion=rotacion,
            costo=costo,
            duracion=duracion,
            solicitar=solicitar,
            ind_duracion=ind_duracion,
            periodo_inicio=fecha_inicio_date,
            periodo_fin=fecha_fin_date,
            periodo_meses=periodo_meses,
        ))

    # --- 5. Guardar todos los detalles en bloque ---
    # Nota: no se usa bulk_update aquí porque carga.detalles.all().delete()
    # ya borró todo lo anterior de esta carga más arriba en la función,
    # así que TODOS los detalles de este período son nuevos.
    with transaction.atomic():
        if detalles_nuevos:
            DetalleInforme.objects.bulk_create(detalles_nuevos, batch_size=500)

    return carga


COLUMNAS_REQUERIDAS_LISTADO = [
    'CODIGO MANTIS', 'DESCRIPCION PRODUCTO', 'LABORATORIO',
    'SALDO DE INVENTARIO', 'COSTO PROMEDIO', 'PRECIO VENTA ANTES IVA',
]
COLUMNAS_REQUERIDAS_VENTAS = ['CODIGO', 'UND']


def validar_estructura_listado(headers):
    """
    Verifica que el archivo de listado básico tenga todas las columnas
    esperadas. Retorna una lista de errores (vacía si todo está bien).
    """
    errores = []
    for columna in COLUMNAS_REQUERIDAS_LISTADO:
        if columna not in headers:
            errores.append(
                f'Falta la columna "{columna}" en el archivo de Listado Básico.'
            )
    return errores


def validar_estructura_ventas(headers):
    """
    Verifica que el archivo de ventas tenga todas las columnas esperadas.
    """
    errores = []
    for columna in COLUMNAS_REQUERIDAS_VENTAS:
        if columna not in headers:
            errores.append(
                f'Falta la columna "{columna}" en el archivo de Ventas.'
            )
    return errores


def leer_fechas_periodo_ventas(archivo_ventas):
    """
    Lee únicamente las celdas de fecha (C2/E2) del archivo de ventas,
    sin parsear todo el contenido, para poder verificar si el período
    ya fue cargado ANTES de procesar el archivo completo.

    Devuelve (fecha_inicio: date, fecha_fin: date) o levanta ValueError.
    """
    archivo_ventas.seek(0)
    contenido = archivo_ventas.read()
    extension = os.path.splitext(archivo_ventas.name)[1].lower()

    if extension == '.xlsx':
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
        ws = wb.active
        fecha_inicio = ws['C2'].value
        fecha_fin = ws['E2'].value
        wb.close()
    elif extension == '.xls':
        wb = xlrd.open_workbook(file_contents=contenido)
        ws = wb.sheet_by_index(0)
        raw_inicio = ws.cell_value(1, 2)
        raw_fin = ws.cell_value(1, 4)
        datemode = wb.datemode
        fecha_inicio = xldate_as_datetime(raw_inicio, datemode) if isinstance(raw_inicio, (int, float)) else raw_inicio
        fecha_fin = xldate_as_datetime(raw_fin, datemode) if isinstance(raw_fin, (int, float)) else raw_fin
    else:
        raise ValueError('Formato de archivo de ventas no soportado. Use .xls o .xlsx')

    archivo_ventas.seek(0)  # dejar el puntero listo para una lectura posterior

    if not isinstance(fecha_inicio, datetime) or not isinstance(fecha_fin, datetime):
        raise ValueError('No se pudieron leer las fechas de inicio y fin del archivo de ventas (celdas C2/E2).')

    return fecha_inicio.date(), fecha_fin.date()


def verificar_periodo_existente(fecha_inicio_date, fecha_fin_date):
    """
    Verifica si ya existe una carga con exactamente ese período.
    Retorna 'nuevo' o 'actualizado', junto con la carga existente si aplica.
    """
    carga_existente = (
        CargaReporte.objects
        .filter(detalles__periodo_inicio=fecha_inicio_date, detalles__periodo_fin=fecha_fin_date)
        .distinct()
        .first()
    )
    if carga_existente:
        return 'actualizado', carga_existente
    return 'nuevo', None


def validar_archivos_antes_de_procesar(archivo_listado, archivo_ventas):
    """
    Validación completa PRE-procesamiento, en memoria, sin tocar la BD
    más allá de una consulta de existencia:
      1. Estructura/columnas de ambos archivos.
      2. Lectura rápida de fechas del archivo de ventas.
      3. Determina si el período es nuevo o ya existe (actualización).

    Retorna un dict:
      {
        'valido': bool,
        'errores': [...],
        'period_status': 'nuevo' | 'actualizado' | None,
        'fecha_inicio': date | None,
        'fecha_fin': date | None,
      }
    No lanza excepciones: toda validación de negocio se reporta en 'errores'.
    """
    errores = []

    try:
        headers_listado, _ = leer_excel(archivo_listado)
        errores += validar_estructura_listado(headers_listado)
    except Exception as e:
        errores.append(f'No se pudo leer el archivo de Listado Básico: {str(e)}')

    fecha_inicio = fecha_fin = None
    try:
        headers_ventas, _ = leer_excel(archivo_ventas, fila_encabezado=3)
        errores += validar_estructura_ventas(headers_ventas)
        fecha_inicio, fecha_fin = leer_fechas_periodo_ventas(archivo_ventas)
    except Exception as e:
        errores.append(f'No se pudo leer el archivo de Ventas: {str(e)}')

    if errores:
        return {
            'valido': False, 'errores': errores, 'period_status': None,
            'fecha_inicio': None, 'fecha_fin': None,
        }

    try:
        calcular_periodo_meses(
            datetime.combine(fecha_inicio, datetime.min.time()),
            datetime.combine(fecha_fin, datetime.min.time()),
        )
    except ValueError as e:
        return {
            'valido': False, 'errores': [str(e)], 'period_status': None,
            'fecha_inicio': None, 'fecha_fin': None,
        }

    period_status, _ = verificar_periodo_existente(fecha_inicio, fecha_fin)

    # Reponer el puntero de ambos archivos para el procesamiento real posterior
    archivo_listado.seek(0)
    archivo_ventas.seek(0)

    return {
        'valido': True, 'errores': [], 'period_status': period_status,
        'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin,
    }


#================================================================
# UTILS PARA PODER EXCLUIR PRODUCTOS DESDE UN FORMULARIO O EXCEL
#================================================================
def leer_excel_exclusiones(ruta_archivo):
    """
    Lee un Excel de productos a excluir. Encabezados en la fila 1, datos desde la fila 2.
    Busca la columna de código bajo varios nombres posibles, y la de observación si existe.
    Devuelve una lista de tuplas (codigo, observacion_o_None).
    """
    extension = os.path.splitext(ruta_archivo)[1].lower()
    posibles_codigo = ['REFERENCIA', 'CODIGO', 'CODIGO MANTIS', 'CÓDIGO']
    posibles_obs = ['EXISTENCIA', 'OBSERVACION', 'OBSERVACIÓN']

    if extension == '.xlsx':
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
        headers = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor:
                headers[str(valor).strip().upper()] = col - 1
        filas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            filas.append([c if c is not None else '' for c in row])
        wb.close()
    elif extension == '.xls':
        wb = xlrd.open_workbook(ruta_archivo)
        ws = wb.sheet_by_index(0)
        headers = {}
        for col in range(ws.ncols):
            valor = ws.cell_value(0, col)
            if valor:
                headers[str(valor).strip().upper()] = col
        filas = []
        for row_idx in range(1, ws.nrows):
            filas.append([ws.cell_value(row_idx, c) for c in range(ws.ncols)])
    else:
        raise ValueError("Formato no soportado. Use archivos .xls o .xlsx")

    col_codigo = next((headers[n] for n in posibles_codigo if n in headers), None)
    if col_codigo is None:
        raise KeyError(
            f"No se encontró columna de código. Columnas disponibles: {list(headers.keys())}"
        )
    col_obs = next((headers[n] for n in posibles_obs if n in headers), None)

    resultados = []
    for fila in filas:
        codigo = str(fila[col_codigo]).strip().upper() if fila[col_codigo] else ''
        if not codigo:
            continue
        observacion = None
        if col_obs is not None and col_obs < len(fila) and fila[col_obs]:
            observacion = str(fila[col_obs]).strip()
        resultados.append((codigo, observacion))
    return resultados


#===========================================================
# UTILS PARA REACTIVAR EL PRODUCTO SI YA ESTABA EXCLUIDO
#===========================================================
@transaction.atomic
def reactivar_producto(exclusion_id):
    """
    Desactiva la exclusión (activo=False) y reactiva el producto (estado=True).
    Devuelve True si se pudo reactivar, False si la exclusión no existía o ya estaba inactiva.
    """
    try:
        exclusion = ProductoExcluido.objects.select_related('producto').get(
            id=exclusion_id, activo=True
        )
    except ProductoExcluido.DoesNotExist:
        return False

    exclusion.activo = False
    exclusion.save(update_fields=['activo'])

    producto = exclusion.producto
    if not producto.estado:
        producto.estado = True
        producto.save(update_fields=['estado'])

    return True

# ===========================================================
# UTILS PARA CARGAR EL DETALLADO DE COMPRA (versión optimizada)
# ===========================================================
COL_FACTURA = column_index_from_string('B')
COL_FECHA = column_index_from_string('C')
COL_CODIGO = column_index_from_string('D')
COL_DESCRIPCION = column_index_from_string('E')
COL_BARRAS = column_index_from_string('G')
COL_LABORATORIO = column_index_from_string('I')
COL_VALOR = column_index_from_string('T')
COL_PROVEEDOR = column_index_from_string('AB')


def _parsear_fecha_compra(valor):
    """
    Acepta datetime nativo de Excel o texto tipo '01/01/26' / '01/01/2026'.
    Nombre específico para evitar colisión con _parsear_fecha_vencimiento.
    """
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str) and valor.strip():
        for fmt in ('%d/%m/%y', '%d/%m/%Y'):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _texto_compra(valor):
    return str(valor).strip() if valor not in (None, '') else ''


def procesar_archivo_compras(archivo):
    """
    Versión optimizada: en vez de una consulta por fila (get_or_create /
    update_or_create), se hace TODO en un puñado de queries:

      1. Se leen todas las filas del Excel a memoria (sin tocar la BD).
      2. Se traen de un solo golpe los productos y compras existentes
         que coinciden con los códigos/facturas del archivo.
      3. Se separan en listas de "nuevos" vs "ya existentes".
      4. Se insertan/actualizan en bloque con bulk_create / bulk_update,
         dentro de una única transacción.

    Esto reduce miles de round-trips de red a ~4-6 queries totales,
    sin cambiar el comportamiento ni el resultado final.
    """
    archivo.seek(0)
    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb.active

    fecha_inicio = _parsear_fecha_compra(ws['B4'].value)
    fecha_fin = _parsear_fecha_compra(ws['D4'].value)

    # --- 1. Leer todas las filas válidas a memoria ---
    filas_validas = []
    codigos_vistos = set()
    for row in ws.iter_rows(min_row=6, values_only=True):
        codigo = _texto_compra(row[COL_CODIGO - 1])
        if not codigo:
            continue

        fecha_compra = _parsear_fecha_compra(row[COL_FECHA - 1])
        if fecha_compra is None:
            continue

        valor_val = row[COL_VALOR - 1]
        filas_validas.append({
            'codigo': codigo,
            'barras': _texto_compra(row[COL_BARRAS - 1]),
            'descripcion': _texto_compra(row[COL_DESCRIPCION - 1]),
            'laboratorio': _texto_compra(row[COL_LABORATORIO - 1]),
            'factura': _texto_compra(row[COL_FACTURA - 1]),
            'proveedor': _texto_compra(row[COL_PROVEEDOR - 1]),
            'valor_compra': float(valor_val) if valor_val else 0,
            'fecha_compra': fecha_compra,
        })
        codigos_vistos.add(codigo)

    wb.close()

    if not filas_validas:
        return 0, 0, []

    # --- 2. Traer productos existentes de un solo golpe ---
    productos_existentes = {
        p.codigo_mantis: p
        for p in Producto.objects.filter(codigo_mantis__in=codigos_vistos)
    }

    productos_nuevos = {}
    productos_creados_codigos = []

    for f in filas_validas:
        codigo = f['codigo']
        if codigo not in productos_existentes and codigo not in productos_nuevos:
            productos_nuevos[codigo] = Producto(
                codigo_mantis=codigo,
                barras=f['barras'],
                descripcion=f['descripcion'],
                laboratorio=f['laboratorio'],
                estado=True,
            )

    # --- 3. Crear productos nuevos en bloque ---
    if productos_nuevos:
        with transaction.atomic():
            Producto.objects.bulk_create(productos_nuevos.values(), batch_size=500)
        # Volver a traerlos para tener sus IDs asignados
        nuevos_creados = Producto.objects.filter(codigo_mantis__in=productos_nuevos.keys())
        for p in nuevos_creados:
            productos_existentes[p.codigo_mantis] = p
            productos_creados_codigos.append(p.codigo_mantis)

    # --- 4. Traer compras existentes de un solo golpe ---
    # Clave natural: (producto_id, factura, fecha_compra)
    facturas_vistas = {f['factura'] for f in filas_validas}
    compras_existentes_qs = Compra.objects.filter(
        producto__codigo_mantis__in=codigos_vistos,
        factura__in=facturas_vistas,
    ).select_related('producto')

    compras_existentes = {
        (c.producto.codigo_mantis, c.factura, c.fecha_compra): c
        for c in compras_existentes_qs
    }

    compras_nuevas = []
    compras_a_actualizar = []
    claves_vistas = set()

    for f in filas_validas:
        clave = (f['codigo'], f['factura'], f['fecha_compra'])
        if clave in claves_vistas:
            # Evita duplicados dentro del mismo archivo (misma fila repetida)
            continue
        claves_vistas.add(clave)

        producto = productos_existentes[f['codigo']]

        if clave in compras_existentes:
            compra = compras_existentes[clave]
            compra.proveedor = f['proveedor']
            compra.valor_compra = f['valor_compra']
            compra.periodo_inicio = fecha_inicio
            compra.periodo_fin = fecha_fin
            compras_a_actualizar.append(compra)
        else:
            compras_nuevas.append(Compra(
                producto=producto,
                factura=f['factura'],
                fecha_compra=f['fecha_compra'],
                proveedor=f['proveedor'],
                valor_compra=f['valor_compra'],
                periodo_inicio=fecha_inicio,
                periodo_fin=fecha_fin,
            ))

    # --- 5. Guardar en bloque dentro de una sola transacción ---
    with transaction.atomic():
        if compras_nuevas:
            Compra.objects.bulk_create(compras_nuevas, batch_size=500)
        if compras_a_actualizar:
            Compra.objects.bulk_update(
                compras_a_actualizar,
                fields=['proveedor', 'valor_compra', 'periodo_inicio', 'periodo_fin'],
                batch_size=500,
            )

    return len(compras_nuevas), len(compras_a_actualizar), productos_creados_codigos


#====================================================================
# UTILS PARA LA VISUALIZACION DE PRODUCTOS EN BAJO STOCK O SIN STOCK 
#====================================================================
def obtener_productos_bajo_duracion(umbral=1.0):
    resultados = cache.get(CACHE_KEY_BAJO_DURACION)
    if resultados is not None:
        return resultados

    ultima_carga = CargaReporte.objects.order_by('-fecha_carga').first()
    if not ultima_carga:
        resultados = []
        cache.set(CACHE_KEY_BAJO_DURACION, resultados, CACHE_TIMEOUT)
        return resultados

    productos_excluidos_ids = ProductoExcluido.objects.filter(
        activo=True
    ).values_list('producto_id', flat=True)

    detalles = (
        ultima_carga.detalles
        .select_related('producto')
        .exclude(producto_id__in=productos_excluidos_ids)
        .filter(duracion__lt=umbral)
        .order_by('producto__codigo_mantis')
    )

    # Traemos TODAS las compras de los productos relevantes en UNA sola query
    producto_ids = [d.producto_id for d in detalles]
    compras_por_producto = {}
    todas_compras = (
        Compra.objects
        .filter(producto_id__in=producto_ids)
        .order_by('producto_id', '-fecha_compra')
    )
    for c in todas_compras:
        lista = compras_por_producto.setdefault(c.producto_id, [])
        if len(lista) < 3:
            lista.append(c)

    resultados = []
    for d in detalles:
        resultados.append({
            'detalle': d,
            'compras': compras_por_producto.get(d.producto_id, [])
        })

    cache.set(CACHE_KEY_BAJO_DURACION, resultados, CACHE_TIMEOUT)
    return resultados

# utils para enviar la alerta

def construir_alertas_baja_duracion(umbral=1.0, solo_categoria_a=True, url_app=""):
    """
    Construye alertas para productos con duración menor al umbral.

    Args:
        umbral (float): duración mínima para considerar bajo stock.
        solo_categoria_a (bool): si True, solo incluye productos categoría 'A'.
        url_app (str): URL de la aplicación para incluir en el retorno.

    Returns:
        tuple: (fecha_referencia, dia_de_corte, alertas, url_app)
    """
    productos = obtener_productos_bajo_duracion(umbral=umbral)

    # Caso sin productos: retornamos inmediatamente con url_app
    if not productos:
        hoy = datetime.now()
        fecha = hoy.strftime("%d/%m/%Y")
        dia = hoy.day
        return fecha, dia, [], url_app

    # Obtener categorizaciones de los productos en una sola consulta
    ids = [item['detalle'].producto_id for item in productos]
    categorizaciones = Categorizacion.objects.filter(producto_id__in=ids)
    mapa_cat = {c.producto_id: c for c in categorizaciones}

    alertas = []
    for item in productos:
        detalle = item['detalle']
        compras = item.get('compras') or []
        cat = mapa_cat.get(detalle.producto_id)

        # Filtro por categoría A
        if solo_categoria_a:
            if not cat or cat.tipo_categoria != 'A':
                continue

        codigo = str(detalle.producto.codigo_mantis)
        nombre = detalle.producto.descripcion or codigo
        duracion = float(detalle.duracion)
        solicitar = float(detalle.solicitar)

        if compras:
            proveedor = compras[0].proveedor or "Sin proveedor"
            fecha_compra = compras[0].fecha_compra.strftime("%d/%m/%Y") if compras[0].fecha_compra else ""
        else:
            proveedor = "Sin proveedor"
            fecha_compra = ""

        categoria = cat.get_tipo_categoria_display() if cat else "Sin categoría"
        analisis = cat.analisis if (cat and cat.analisis) else ""

        mensaje = (
            f"Duración de {duracion} (categoría {categoria}). "
            f"Cantidad sugerida a solicitar: {solicitar}."
        )

        alertas.append({
            "codigo": codigo,
            "nombre": nombre,
            "duracion": duracion,
            "solicitar": solicitar,
            "proveedor": proveedor,
            "categoria": categoria,
            "analisis": analisis,
            "fecha_compra": fecha_compra,
            "mensaje": mensaje,
        })

    # Fecha de referencia basada en la última carga
    ultima_carga = CargaReporte.objects.order_by('-fecha_carga').first()
    if ultima_carga:
        fecha_referencia = ultima_carga.fecha_carga.strftime("%d/%m/%Y")
        dia_de_corte = ultima_carga.fecha_carga.day
    else:
        hoy = datetime.now()
        fecha_referencia = hoy.strftime("%d/%m/%Y")
        dia_de_corte = hoy.day

    return fecha_referencia, dia_de_corte, alertas, url_app

#=======================================
# utils para la categorizacion
#=======================================
def convertir(valor):

    if valor is None:
        return 0

    try:
        return float(valor)
    except:
        return 0


def _contar_celdas_rellenas(fila):
    return sum(1 for celda in fila if celda.value not in (None, ""))


def _extraer_categoria_letra(valor):
    """
    Convierte 'Categoría C' -> 'C'. Si ya viene como 'C', la deja igual.
    """
    texto = str(valor).strip().upper()
    if texto.startswith("CATEGORÍA") or texto.startswith("CATEGORIA"):
        return texto.split()[-1]
    return texto


def _leer_filas_formato_viejo(ws):
    """
    Devuelve una lista de tuplas (codigo, categoria, analisis) ya parseadas,
    sin tocar la base de datos.
    """
    registros = []
    for fila in ws.iter_rows(min_row=3):

        categoria = fila[2].value          # C
        codigo = fila[3].value             # D

        valor_pm = convertir(fila[8].value)        # I
        valor_saludia = convertir(fila[10].value)  # K

        if not codigo:
            continue

        codigo = str(codigo).strip()
        categoria = str(categoria).strip().upper()
        analisis = "vende mas PM" if valor_pm > valor_saludia else "vende mas Saludia"

        registros.append((codigo, categoria, analisis))
    return registros


def _importar_categorizacion_formato_viejo(ws):
    registros = _leer_filas_formato_viejo(ws)
    return _guardar_categorizaciones_bulk(registros)


def _leer_filas_formato_nuevo(ws):
    """
    Col A -> Código Mantis
    Col C -> Categoría (ej: 'Categoría C')
    Col D -> Análisis
    Encabezado en fila 1, datos desde fila 2.
    """
    registros = []
    for fila in ws.iter_rows(min_row=2):

        codigo = fila[0].value       # A
        categoria = fila[2].value    # C
        analisis = fila[3].value     # D

        if not codigo:
            continue

        codigo = str(codigo).strip()
        categoria = _extraer_categoria_letra(categoria) if categoria else ""
        analisis = str(analisis).strip() if analisis else ""

        registros.append((codigo, categoria, analisis))
    return registros


def _importar_categorizacion_formato_nuevo(ws):
    registros = _leer_filas_formato_nuevo(ws)
    return _guardar_categorizaciones_bulk(registros)


def _guardar_categorizaciones_bulk(registros):
    """
    registros: lista de tuplas (codigo, categoria, analisis)
    Hace el guardado en batch: 1 query para traer productos,
    1 query para traer categorizaciones existentes,
    1 bulk_create y 1 bulk_update.
    """
    codigos = [r[0] for r in registros]

    productos = Producto.objects.filter(codigo_mantis__in=codigos)
    productos_por_codigo = {p.codigo_mantis: p for p in productos}

    no_encontrados = 0
    validos = []
    for codigo, categoria, analisis in registros:
        producto = productos_por_codigo.get(codigo)
        if not producto:
            no_encontrados += 1
            continue
        validos.append((producto, categoria, analisis))

    producto_ids = [p.pk for p, _, _ in validos]
    existentes = Categorizacion.objects.filter(producto_id__in=producto_ids)
    existentes_por_producto = {c.producto_id: c for c in existentes}

    a_crear = []
    a_actualizar = []

    for producto, categoria, analisis in validos:
        existente = existentes_por_producto.get(producto.pk)
        if existente:
            existente.tipo_categoria = categoria
            existente.analisis = analisis
            a_actualizar.append(existente)
        else:
            a_crear.append(Categorizacion(
                producto=producto,
                tipo_categoria=categoria,
                analisis=analisis,
            ))

    with transaction.atomic():
        if a_crear:
            Categorizacion.objects.bulk_create(a_crear)
        if a_actualizar:
            Categorizacion.objects.bulk_update(a_actualizar, ["tipo_categoria", "analisis"])

    return len(a_crear), len(a_actualizar), no_encontrados


def importar_categorizacion_excel(archivo):

    wb = load_workbook(archivo, data_only=True)
    ws = wb.active

    if ws.max_column > 5:
        creados, actualizados, no_encontrados = _importar_categorizacion_formato_viejo(ws)
    else:
        creados, actualizados, no_encontrados = _importar_categorizacion_formato_nuevo(ws)

    return (
        f"Creados: {creados} | "
        f"Actualizados: {actualizados} | "
        f"No encontrados: {no_encontrados}"
    )

#====================================================
# Utils para la comparacion de reportes de productos 
#====================================================
def obtener_carga_anterior(carga_actual):
    """Retorna la CargaReporte inmediatamente anterior a la actual, o None."""
    return (
        CargaReporte.objects
        .filter(fecha_carga__lt=carga_actual.fecha_carga)
        .order_by('-fecha_carga')
        .first()
    )


def comparar_producto(detalle_actual, carga_anterior):
    """
    Compara la rotación de un producto contra su registro en la carga
    anterior. Retorna un dict con el estado de la comparación.
    """
    detalle_anterior = None
    if carga_anterior:
        detalle_anterior = (
            DetalleInforme.objects
            .filter(carga=carga_anterior, producto=detalle_actual.producto)
            .first()
        )

    rot_actual = float(detalle_actual.rotacion)

    if not detalle_anterior:
        return {
            'rotacion_actual': rot_actual,
            'rotacion_anterior': None,
            'diferencia': None,
            'porcentaje': None,
            'estado': 'SIN_DATOS_ANTERIORES',
        }

    rot_anterior = float(detalle_anterior.rotacion)
    diferencia = rot_actual - rot_anterior

    if rot_anterior == 0:
        porcentaje = None
    else:
        porcentaje = (diferencia / rot_anterior) * 100

    if diferencia > 0:
        estado = 'AUMENTO'
    elif diferencia < 0:
        estado = 'DISMINUCION'
    else:
        estado = 'IGUAL'

    return {
        'rotacion_actual': rot_actual,
        'rotacion_anterior': rot_anterior,
        'diferencia': round(diferencia, 3),
        'porcentaje': round(porcentaje, 1) if porcentaje is not None else None,
        'estado': estado,
        'periodo_anterior': {
            'inicio': detalle_anterior.periodo_inicio.strftime('%d/%m/%Y') if detalle_anterior.periodo_inicio else None,
            'fin': detalle_anterior.periodo_fin.strftime('%d/%m/%Y') if detalle_anterior.periodo_fin else None,
        }
    }


def obtener_comparativa_general(carga_actual, productos_excluidos_ids):
    carga_anterior = obtener_carga_anterior(carga_actual)

    detalles_actual = (
        carga_actual.detalles
        .select_related('producto')
        .exclude(producto_id__in=productos_excluidos_ids)
    )

    prev_map = {}
    if carga_anterior:
        prev_map = {
            d.producto_id: d
            for d in carga_anterior.detalles.select_related('producto').all()
        }

    comparativas = []
    for d in detalles_actual:
        prev = prev_map.get(d.producto_id)
        rot_actual = float(d.rotacion)

        if prev is None:
            continue

        rot_anterior = float(prev.rotacion)
        diferencia = rot_actual - rot_anterior
        porcentaje = (diferencia / rot_anterior * 100) if rot_anterior != 0 else None

        comparativas.append({
            'codigo': d.producto.codigo_mantis,
            'descripcion': d.producto.descripcion,
            'rotacion_actual': rot_actual,
            'rotacion_anterior': rot_anterior,
            'diferencia': round(diferencia, 3),
            'porcentaje': round(porcentaje, 1) if porcentaje is not None else None,
            'ind_duracion_actual': d.ind_duracion,
            'detalle': d,
        })

    suben = sorted(comparativas, key=lambda x: x['diferencia'], reverse=True)[:10]
    candidatos_bajan = [c for c in comparativas if c['ind_duracion_actual'] != 'SIN ROTACIÓN']
    bajan = sorted(candidatos_bajan, key=lambda x: x['diferencia'])[:10]

    periodo_inicio = detalles_actual.first().periodo_inicio if detalles_actual.exists() else None
    periodo_fin = detalles_actual.first().periodo_fin if detalles_actual.exists() else None

    # Umbral (en días) para considerar una compra "reciente" respecto al
    # cierre del período de ventas analizado.
    UMBRAL_DIAS_COMPRA_RECIENTE = 15

    rotacion_baja_compra = []
    if periodo_inicio and periodo_fin:
        candidatos_baja = [
            c for c in comparativas
            if c['ind_duracion_actual'] in ('SIN ROTACIÓN', 'BAJO STOCK')
        ]

        # Una sola query para TODOS los candidatos, en vez de una por producto
        producto_ids_candidatos = [c['detalle'].producto_id for c in candidatos_baja]
        compras_recientes_qs = (
            Compra.objects
            .filter(
                producto_id__in=producto_ids_candidatos,
                fecha_compra__range=[periodo_inicio, periodo_fin],
            )
            .order_by('producto_id', '-fecha_compra')
        )
        # Como viene ordenado por producto_id y fecha desc, la primera
        # aparición de cada producto_id ya es su compra más reciente.
        ultima_compra_por_producto = {}
        for compra in compras_recientes_qs:
            if compra.producto_id not in ultima_compra_por_producto:
                ultima_compra_por_producto[compra.producto_id] = compra

        for c in candidatos_baja:
            d = c['detalle']
            ultima_compra = ultima_compra_por_producto.get(d.producto_id)

            if not ultima_compra:
                continue

            dias_desde_compra = (periodo_fin - ultima_compra.fecha_compra).days
            compra_reciente = dias_desde_compra <= UMBRAL_DIAS_COMPRA_RECIENTE

            rotacion_baja_compra.append({
                'codigo': c['codigo'],
                'descripcion': c['descripcion'],
                'rotacion_actual': c['rotacion_actual'],
                'rotacion_anterior': c['rotacion_anterior'],
                'diferencia': c['diferencia'],
                'porcentaje': c['porcentaje'],
                'ind_duracion_actual': c['ind_duracion_actual'],
                'fecha_compra': ultima_compra.fecha_compra.strftime('%d/%m/%Y'),
                'dias_desde_compra': dias_desde_compra,
                'compra_reciente': compra_reciente,
            })

    def limpiar(lista):
        return [{k: v for k, v in item.items() if k != 'detalle'} for item in lista]

    return {
        'suben': limpiar(suben),
        'bajan': limpiar(bajan),
        'rotacion_baja_compra': rotacion_baja_compra,
        'hay_periodo_anterior': carga_anterior is not None,
    }

#======================================
# CREACION DE UN PRODUCTO
#======================================

def crear_producto(data):
    """
    Crea un producto y retorna un diccionario con el resultado.
    """

    codigo_mantis = data.get("codigo_mantis")

    if Producto.objects.filter(codigo_mantis=codigo_mantis).exists():
        return {
            "success": False,
            "message": "Ya existe un producto con ese código Mantis."
        }

    producto = Producto.objects.create(
        codigo_mantis=codigo_mantis,
        barras=data.get("barras"),
        descripcion=data.get("descripcion"),
        laboratorio=data.get("laboratorio"),
        estado=data.get("estado", True)
    )

    return {
        "success": True,
        "message": "Producto creado correctamente.",
        "producto": producto
    }

# actualizar producto
def actualizar_producto(id_random, data):
    try:
        producto = Producto.objects.get(id_random=id_random)
    except Producto.DoesNotExist:
        return {
            "success": False,
            "message": "El producto no existe."
        }

    codigo_mantis = data.get("codigo_mantis")

    if Producto.objects.filter(codigo_mantis=codigo_mantis).exclude(id_random=id_random).exists():
        return {
            "success": False,
            "message": "Ya existe otro producto con ese código Mantis."
        }

    producto.codigo_mantis = codigo_mantis
    producto.barras = data.get("barras")
    producto.descripcion = data.get("descripcion")
    producto.laboratorio = data.get("laboratorio")
    producto.estado = data.get("estado", True)

    producto.save()

    return {
        "success": True,
        "message": "Producto actualizado correctamente.",
        "producto": producto
    }

#=======================================
# UTILS PARA LAS FECHAS DE VENCIMIENTO
#=======================================
def _limpiar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


# ─────────────────────────────────────────────
# CAPA DE ABSTRACCIÓN: LEER XLSX O XLS
# Devuelve una lista de filas, cada fila es una lista de valores crudos
# (igual formato sin importar si el archivo es .xlsx o .xls)
# ─────────────────────────────────────────────
def _leer_filas_excel(archivo):
    """
    Detecta la extensión del archivo y devuelve una lista de filas
    (cada fila es una lista de valores), normalizando fechas de Excel
    (números seriales) a objetos date cuando corresponde.
    """
    nombre = archivo.name.lower()

    if nombre.endswith(".xlsx"):
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
        filas = []
        for fila in ws.iter_rows():
            filas.append([celda.value for celda in fila])
        return filas

    elif nombre.endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=archivo.read())
        ws = wb.sheet_by_index(0)
        filas = []
        for num_fila in range(ws.nrows):
            fila_valores = []
            for num_col in range(ws.ncols):
                celda = ws.cell(num_fila, num_col)
                valor = celda.value
                # xlrd guarda fechas como número serial (tipo XL_CELL_DATE = 3)
                if celda.ctype == 3 and valor:
                    try:
                        tupla_fecha = xlrd.xldate_as_tuple(valor, wb.datemode)
                        valor = datetime(*tupla_fecha)
                    except Exception:
                        pass
                fila_valores.append(valor)
            filas.append(fila_valores)
        return filas

    else:
        raise ValueError("Formato de archivo no soportado. Solo se aceptan .xlsx y .xls")


# ─────────────────────────────────────────────
# EXTRACCIÓN POR TIPO DE ARCHIVO
# (ahora trabajan sobre la lista de filas ya normalizada, no sobre ws)
# ─────────────────────────────────────────────
def _extraer_fechas_vencimiento(filas):
    """
    Col A (índice 0)  -> Código Artículo  (encabezado fila 3 -> índice 2, datos desde fila 4 -> índice 3)
    Col K (índice 10) -> Lote
    Col L (índice 11) -> Lote Vencimiento
    Col M (índice 12) -> Unidades
    """
    registros = []
    for fila in filas[3:]:  # datos desde la fila 4 (índice 3)
        if len(fila) <= 12:
            continue
        codigo = _limpiar_texto(fila[0])
        lote = _limpiar_texto(fila[10])
        fecha_venc = _parsear_fecha(fila[11])
        unidades = fila[12] if fila[12] not in (None, "") else None

        if not codigo or not lote or not fecha_venc:
            continue

        registros.append({
            "codigo": codigo,
            "lote": lote,
            "fecha_vencimiento": fecha_venc,
            "unidades": unidades,
        })
    return registros


# ─────────────────────────────────────────────
# IMPORTACIÓN PRINCIPAL
# ─────────────────────────────────────────────
def importar_vencimientos_excel(archivo):
    """
    Formato único de archivo (Codigo Articulo, Lote, Lote Vencimiento).
    Soporta archivos .xlsx y .xls indistintamente.

    Versión optimizada: en vez de 2 queries por fila (.get() del producto
    + .get_or_create() del vencimiento), se hace todo en un puñado de
    consultas totales, sin importar cuántas filas tenga el archivo.
    """
    try:
        filas = _leer_filas_excel(archivo)
    except ValueError as e:
        return str(e), []

    registros = _extraer_fechas_vencimiento(filas)

    if not registros:
        return "No se encontraron registros válidos en el archivo.", []

    # --- 1. Traer todos los productos involucrados en una sola consulta ---
    codigos = {reg["codigo"] for reg in registros}
    productos_por_codigo = {
        p.codigo_mantis: p
        for p in Producto.objects.filter(codigo_mantis__in=codigos)
    }

    no_encontrados = []
    validos = []
    for reg in registros:
        producto = productos_por_codigo.get(reg["codigo"])
        if not producto:
            no_encontrados.append(reg["codigo"])
            continue
        validos.append((producto, reg["lote"], reg["fecha_vencimiento"], reg.get("unidades")))

    if not validos:
        mensaje = (
            f"Creados: 0 | Ya existían (omitidos): 0 | "
            f"Productos no encontrados: {len(no_encontrados)}"
        )
        return mensaje, no_encontrados

    # --- 2. Traer los vencimientos ya existentes (producto, lote) en una sola consulta ---
    producto_ids = {p.pk for p, _, _, _ in validos}
    existentes = set(
        Vencimiento.objects
        .filter(producto_id__in=producto_ids)
        .values_list("producto_id", "lote")
    )

    # --- 3. Separar en "a crear" y "ya existían", evitando duplicados dentro del mismo archivo ---
    a_crear = []
    claves_vistas = set()
    ya_existian = 0

    for producto, lote, fecha_venc, unidades in validos:
        clave = (producto.pk, lote)

        if clave in existentes or clave in claves_vistas:
            ya_existian += 1
            continue

        claves_vistas.add(clave)
        a_crear.append(Vencimiento(
            producto=producto,
            lote=lote,
            fecha_vencimiento=fecha_venc,
            unidades=unidades,
        ))

    # --- 4. Guardar en bloque dentro de una sola transacción ---
    with transaction.atomic():
        if a_crear:
            Vencimiento.objects.bulk_create(a_crear, batch_size=500)

    mensaje = (
        f"Creados: {len(a_crear)} | "
        f"Ya existían (omitidos): {ya_existian} | "
        f"Productos no encontrados: {len(no_encontrados)}"
    )

    return mensaje, no_encontrados