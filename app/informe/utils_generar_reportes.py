"""
======================================================================
UTILIDADES PARA GENERAR EL REPORTE DE INVENTARIO EN EXCEL
======================================================================
Genera un archivo .xlsx con el mismo layout que la plantilla original:

    Fila 1  -> "MESES A REVISAR:" (A1) y número de meses (C1)
    Fila 3  -> Encabezados
    Fila 4+ -> Datos (uno por producto / DetalleInforme)

Columnas:
    A CODIGO MANTIS
    B BARRAS
    C DESCRIPCION
    D LABORATORIO
    E STOCK
    F RENTABILIDAD DESC   (formato %)
    G ROTACION
    H COSTO               (formato moneda)
    I DURACION POR MES
    J OBSERVACIÓN
    K SOLICITAR
    L PEDIDOS REALIZADOS
    M IND. DURACION        <- coloreada según el estado

Colores del indicador (columna M, fondo y texto) basados en el template:
    SOBRE STOCK  -> fondo #cce5f6, texto #2471a3
    OK           -> fondo #d5f5e3, texto #1e8449
    BAJO STOCK   -> fondo #ffe2bf, texto #d35400
    SIN STOCK    -> fondo #f4dbd8, texto #c0392b
    SIN ROTACIÓN -> fondo #e4e8e9, texto #5d6d7e

Además, se agrega un AutoFilter sobre el encabezado (fila 3) y formato
condicional para que los colores se mantengan al filtrar u ordenar.
======================================================================
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule
from io import BytesIO
from openpyxl import Workbook, load_workbook
from datetime import datetime
from django.http import HttpResponse
from openpyxl.worksheet.datavalidation import DataValidation
from .models import *
from django.contrib.staticfiles import finders



# ---------------------------------------------------------------------
# Colores por estado (ARGB). Basados en el template, con fondo pastel
# y texto en color oscuro específico para cada estado.
# ---------------------------------------------------------------------
COLORES_ESTADO = {
    'SOBRE STOCK': {
        'fill': 'FFCCE5F6',   # fondo azul pastel
        'font': Font(color='FF2471A3', bold=True)
    },
    'OK': {
        'fill': 'FFD5F5E3',   # fondo verde pastel
        'font': Font(color='FF1E8449', bold=True)
    },
    'BAJO STOCK': {
        'fill': 'FFFFE2BF',   # fondo naranja pastel
        'font': Font(color='FFD35400', bold=True)
    },
    'SIN STOCK': {
        'fill': 'FFF4DBD8',   # fondo rojo pastel
        'font': Font(color='FFC0392B', bold=True)
    },
    'SIN ROTACIÓN': {
        'fill': 'FFE4E8E9',   # fondo gris pastel
        'font': Font(color='FF5D6D7E', bold=True)
    },
}

FUENTE_NEGRA = Font(color='FF000000', bold=True)

ENCABEZADOS = [
    'CODIGO MANTIS', 'BARRAS', 'DESCRIPCION', 'LABORATORIO', 'STOCK',
    'RENTABILIDAD DESC', 'ROTACION ', 'COSTO', 'DURACION POR MES',
    'OBSERVACIÓN ', 'SOLICITAR', 'PEDIDOS REALIZADOS', 'IND. DURACION',
]

FORMATOS_NUMERO = {
    'B': '0',
    'E': 'General',
    'F': '0%',
    'G': '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-',
    'H': '_-"$"\\ * #,##0.00_-;\\-"$"\\ * #,##0.00_-;_-"$"\\ * "-"??_-;_-@_-',
    'I': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"??_-;_-@_-',
    'K': '_-* #,##0.0_-;\\-* #,##0.0_-;_-* "-"??_-;_-@_-',
    'L': '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-',
}

BORDE_FINO = Border(
    left=Side(style='thin', color='FF808080'),
    right=Side(style='thin', color='FF808080'),
    top=Side(style='thin', color='FF808080'),
    bottom=Side(style='thin', color='FF808080'),
)

RELLENO_BLANCO = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')


def _aplicar_formato_condicional_estado(ws, col_letra, primera_fila, ultima_fila):
    """
    Agrega reglas de formato condicional en la columna de estado
    (IND. DURACION) para que, sin importar cómo el usuario filtre u
    ordene la tabla, el color siga apareciendo automáticamente según el
    texto de la celda.
    """
    rango = f'{col_letra}{primera_fila}:{col_letra}{ultima_fila}'
    for estado, config in COLORES_ESTADO.items():
        ws.conditional_formatting.add(
            rango,
            CellIsRule(
                operator='equal',
                formula=[f'"{estado}"'],
                fill=PatternFill(start_color=config['fill'], end_color=config['fill'], fill_type='solid'),
                font=config['font'],
            ),
        )


def generar_reporte_excel(detalles, meses_a_revisar, ruta_salida=None):
    """
    Genera el reporte de inventario en Excel a partir de un queryset (o
    lista) de instancias de DetalleInforme (select_related('producto')
    recomendado por rendimiento).

    Parámetros
    ----------
    detalles : iterable de DetalleInforme
    meses_a_revisar : int
        Número de meses que cubre el período analizado (se muestra en
        C1, igual que en la plantilla original).
    ruta_salida : str | None
        Si se indica, el archivo se guarda en esa ruta. Si es None, se
        retorna un BytesIO listo para usarse en un HttpResponse.

    Retorna
    -------
    str  -> ruta del archivo, si se indicó ruta_salida
    BytesIO -> buffer del archivo, si ruta_salida es None
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hoja1'

    # ---------------------------------------------------------
    # Fila 1: Meses a revisar
    # ---------------------------------------------------------
    ws['A1'] = 'MESES A REVISAR:'
    ws['A1'].font = Font(bold=True)
    ws['B1'] = meses_a_revisar
    ws['B1'].font = Font(bold=True)

    # ---------------------------------------------------------
    # Fila 3: Encabezados
    # ---------------------------------------------------------
    fila_encabezado = 3
    relleno_encabezado = PatternFill(start_color='FF203864', end_color='FF203864', fill_type='solid')
    for idx, titulo in enumerate(ENCABEZADOS, start=1):
        celda = ws.cell(row=fila_encabezado, column=idx, value=titulo)
        celda.font = Font(bold=True, color='FFFFFFFF')
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border = BORDE_FINO

    # ---------------------------------------------------------
    # Filas de datos
    # ---------------------------------------------------------
    fila = fila_encabezado + 1
    primera_fila_datos = fila

    for d in detalles:
        producto = d.producto
        ws.cell(row=fila, column=1, value=producto.codigo_mantis)
        ws.cell(row=fila, column=2, value=producto.barras)
        ws.cell(row=fila, column=3, value=producto.descripcion)
        ws.cell(row=fila, column=4, value=producto.laboratorio)
        ws.cell(row=fila, column=5, value=float(d.stock))
        ws.cell(row=fila, column=6, value=float(d.rentabilidad) / 100)  # % almacenado como fracción
        ws.cell(row=fila, column=7, value=float(d.rotacion))
        ws.cell(row=fila, column=8, value=float(d.costo))
        ws.cell(row=fila, column=9, value=float(d.duracion))
        ws.cell(row=fila, column=10, value=d.observacion or '')
        ws.cell(row=fila, column=11, value=float(d.solicitar))
        ws.cell(row=fila, column=12, value=float(d.pedidos) if d.pedidos else '')
        celda_estado = ws.cell(row=fila, column=13, value=d.ind_duracion)

        # Formatos numéricos por columna
        for col_letra, formato in FORMATOS_NUMERO.items():
            ws[f'{col_letra}{fila}'].number_format = formato

        # Bordes, fondo blanco y alineación general de la fila
        for col in range(1, 14):
            c = ws.cell(row=fila, column=col)
            c.border = BORDE_FINO
            c.fill = RELLENO_BLANCO
            if col in (3,):  # descripción alineada a la izquierda
                c.alignment = Alignment(horizontal='left', vertical='center')
            else:
                c.alignment = Alignment(horizontal='center', vertical='center')

        # Aplicar color de fondo y fuente a la celda de estado según el valor
        estado = (d.ind_duracion or '').strip().upper()
        config = COLORES_ESTADO.get(estado)
        if config:
            celda_estado.fill = PatternFill(start_color=config['fill'], end_color=config['fill'], fill_type='solid')
            celda_estado.font = config['font']

        fila += 1

    ultima_fila_datos = fila - 1

    # Si no hubo datos, dejamos al menos el encabezado
    if ultima_fila_datos < primera_fila_datos:
        ultima_fila_datos = primera_fila_datos

    # ---------------------------------------------------------
    # Formato condicional dinámico sobre la columna de estado
    # ---------------------------------------------------------
    _aplicar_formato_condicional_estado(ws, 'M', primera_fila_datos, ultima_fila_datos)

    # ---------------------------------------------------------
    # Tabla con autofiltro
    # ---------------------------------------------------------
    rango_tabla = f'A{fila_encabezado}:M{ultima_fila_datos}'
    tabla = Table(displayName='TablaInventario', ref=rango_tabla)
    estilo = TableStyleInfo(
        name='TableStyleLight1',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    # Reforzar bordes y fondo blanco después de la tabla
    for r in range(primera_fila_datos, ultima_fila_datos + 1):
        for col in range(1, 14):
            c = ws.cell(row=r, column=col)
            c.border = BORDE_FINO

    # ---------------------------------------------------------
    # Ancho de columnas
    # ---------------------------------------------------------
    anchos = {
        'A': 14, 'B': 16, 'C': 45, 'D': 18, 'E': 10, 'F': 16,
        'G': 12, 'H': 14, 'I': 16, 'J': 20, 'K': 12, 'L': 18, 'M': 16,
    }
    for col_letra, ancho in anchos.items():
        ws.column_dimensions[col_letra].width = ancho

    ws.freeze_panes = 'A4'
    ws.row_dimensions[fila_encabezado].height = 30

    # ---------------------------------------------------------
    # Leyenda de colores (debajo de la tabla)
    # ---------------------------------------------------------
    fila_leyenda = ultima_fila_datos + 3
    ws.cell(row=fila_leyenda, column=1, value='LEYENDA IND. DURACION:').font = Font(bold=True)
    fila_leyenda += 1
    for estado, config in COLORES_ESTADO.items():
        # Mostrar todos los estados, incluido SIN ROTACIÓN
        c_color = ws.cell(row=fila_leyenda, column=1, value='   ')
        c_color.fill = PatternFill(start_color=config['fill'], end_color=config['fill'], fill_type='solid')
        ws.cell(row=fila_leyenda, column=2, value=estado).font = config['font']
        fila_leyenda += 1

    # ---------------------------------------------------------
    # Guardar / retornar
    # ---------------------------------------------------------
    if ruta_salida:
        wb.save(ruta_salida)
        return ruta_salida

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

#================================================
# utils para descargar el reporte de compras
#================================================
def preparar_datos_productos_con_compras(productos):
    """
    Toma la lista de productos devuelta por obtener_productos_bajo_duracion
    y devuelve una lista de diccionarios, uno por cada compra asociada
    (incluyendo la más reciente y las anteriores).
    """
    datos = []
    for item in productos:
        detalle = item['detalle']
        compras = item.get('compras', [])  # lista de objetos Compra
        categorizacion = item.get('categorizacion')

        if not compras:
            # Si no hay compras, igual agregamos una fila con datos vacíos
            datos.append({
                'codigo': detalle.producto.codigo_mantis,
                'nombre': detalle.producto.descripcion,
                'duracion': detalle.duracion,
                'solicitar': detalle.solicitar,
                'proveedor': '',
                'categoria': categorizacion.get_tipo_categoria_display() if categorizacion else '',
                'analisis': categorizacion.analisis if categorizacion else '',
                'fecha_compra': None,
                'costo': detalle.costo,
            })
        else:
            # Para cada compra (ordenadas por fecha desc según tu consulta)
            for idx, compra in enumerate(compras):
                datos.append({
                    'codigo': detalle.producto.codigo_mantis,
                    'nombre': detalle.producto.descripcion,
                    'duracion': detalle.duracion,
                    'solicitar': detalle.solicitar,
                    'proveedor': compra.proveedor,
                    'categoria': categorizacion.get_tipo_categoria_display() if categorizacion else '',
                    'analisis': categorizacion.analisis if categorizacion else '',
                    'fecha_compra': compra.fecha_compra,
                    'costo': detalle.costo,
                })
    return datos


def generar_excel_reporte(productos, titulo="Productos con baja duración"):
    """
    Genera un libro con:
      - Hoja "Detalle": listado con columna "Seleccionar" (checkbox Sí/No)
      - Hoja "Orden de Compra": estructura idéntica a ORDEN_DE_COMPRA.xlsx
        (3 filas de cabecera + datos), que el macro VBA llena cuando
        se marca el checkbox en Hoja1.
    IMPORTANTE: para que el checkbox dispare la macro, este archivo debe
    partir de la plantilla .xlsm que ya tiene el código VBA (ver punto 2).
    """
    datos = preparar_datos_productos_con_compras(productos)

    # --- Localizar la plantilla macro-habilitada mediante finders ---
    ruta_plantilla = finders.find("plantillas/plantilla_reporte_base.xlsm")
    if not ruta_plantilla:
        raise FileNotFoundError(
            "No se encontró la plantilla plantillas/plantilla_reporte_base.xlsm. "
            "Verifica que esté en tu carpeta static/ y que STATICFILES_DIRS/STATIC_ROOT "
            "estén configurados correctamente."
        )
    wb = load_workbook(ruta_plantilla, keep_vba=True)

    ws1 = wb["Detalle"]
    ws2 = wb["OrdenCompra"]  # ya trae la estructura de 3 filas de cabecera

    # Limpiar filas de datos previas (si la plantilla trae ejemplos)
    ws1.delete_rows(1, ws1.max_row)
    if ws2.max_row > 3:
        ws2.delete_rows(4, ws2.max_row - 3)

    # --- Encabezados Hoja1 ---
    headers = [
        "Código", "Nombre", "Duración", "Cantidad a Solicitar",
        "Proveedor", "Categoría", "Análisis", "Fecha de Compra", "Último Costo","Seleccionar"

    ]
    ws1.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Validación de datos: lista Sí/No como "checkbox" ---
    dv = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    ws1.add_data_validation(dv)

    # DESPUÉS
    for d in datos:
        row = [
            d['codigo'], d['nombre'], d['duracion'], d['solicitar'],
            d['proveedor'], d['categoria'], d['analisis'],
            d['fecha_compra'].strftime("%d/%m/%Y") if d['fecha_compra'] else "",
            d['costo'],
            "No",
        ]
        ws1.append(row)
        dv.add(ws1.cell(row=ws1.max_row, column=10))  # columna J (10), "Seleccionar" — la macro depende de esto

    ws1.auto_filter.ref = f"A1:J{ws1.max_row}"  # rango hasta columna J (ahora 10 columnas)

    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws1.column_dimensions[column].width = min(max_length + 2, 40)

    # Guardar en memoria con BytesIO en lugar de disco
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        content=buffer.getvalue(),
        content_type='application/vnd.ms-excel.sheet.macroEnabled.12'
    )
    filename = f"reporte_baja_duracion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsm"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

#===============================================
# GENERAR REPORTE DE PRODUCTOS EXCLUIDOS
#===============================================
def generar_reporte_productos_excluidos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Excluidos"

    encabezados = [
        "Código Mantis",
        "Descripción",
        "Observación",
        "Fecha Exclusión",
    ]

    for col, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = titulo
        celda.font = Font(bold=True)

    # SOLO PRODUCTOS ACTUALMENTE EXCLUIDOS
    exclusiones = (
        ProductoExcluido.objects
        .filter(activo=True)
        .select_related("producto")
        .order_by("-fecha_creacion")
    )

    fila = 2

    for exclusion in exclusiones:
        ws.cell(fila, 1).value = exclusion.producto.codigo_mantis
        ws.cell(fila, 2).value = exclusion.producto.descripcion
        ws.cell(fila, 3).value = exclusion.observacion or ""
        ws.cell(fila, 4).value = exclusion.fecha_creacion.strftime("%d/%m/%Y %H:%M")
        fila += 1

    # Ajustar ancho automáticamente
    for columna in ws.columns:
        ancho = max(len(str(celda.value or "")) for celda in columna)
        ws.column_dimensions[columna[0].column_letter].width = ancho + 5

    return wb
# utils para formatear unidades de manera consistente en los reportes de vencimientos y categorizaciones.
def _formatear_unidades(valor):
    if valor is None:
        return None
    if valor == valor.to_integral_value():
        return int(valor)
    return float(valor)

#================================================
# REPORTE DE FECHAS DE VENCIMIENTO 
#================================================
def generar_reporte_vencimientos(fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vencimientos"

    encabezados = [
        "Código Mantis",
        "Descripción",
        "Lote",
        "Unidades",
        "Fecha Vencimiento",
        "Fecha Carga",
    ]

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True)

    vencimientos = (
        Vencimiento.objects
        .select_related("producto")
        .filter(
            fecha_vencimiento__range=[fecha_inicio, fecha_fin]
        )
        .order_by("fecha_vencimiento")
    )

    fila = 2

    for item in vencimientos:
        ws.cell(fila, 1).value = item.producto.codigo_mantis
        ws.cell(fila, 2).value = item.producto.descripcion
        ws.cell(fila, 3).value = item.lote
        ws.cell(fila, 4).value = _formatear_unidades(item.unidades)
        ws.cell(fila, 5).value = item.fecha_vencimiento.strftime("%d/%m/%Y")
        ws.cell(fila, 6).value = item.fecha_carga.strftime("%d/%m/%Y %H:%M")
        fila += 1

    # Ajustar ancho automáticamente
    for columna in ws.columns:
        longitud = max(len(str(celda.value or "")) for celda in columna)
        ws.column_dimensions[columna[0].column_letter].width = longitud + 5

    return wb

#========================================
# UTILS PARA GENERAR REPORTE DE CATEGORIZACION
#===========================================
def generar_reporte_categorizaciones():
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorizaciones"

    encabezados = [
        "Código Mantis",
        "Descripción",
        "Categoría",
        "Análisis",
        "Fecha Registro",
    ]

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col)
        celda.value = encabezado
        celda.font = Font(bold=True)

    categorizaciones = (
        Categorizacion.objects
        .select_related("producto")
        .order_by("-fecha_registro")
    )

    fila = 2

    for categoria in categorizaciones:
        ws.cell(fila, 1).value = categoria.producto.codigo_mantis
        ws.cell(fila, 2).value = categoria.producto.descripcion
        ws.cell(fila, 3).value = categoria.get_tipo_categoria_display()
        ws.cell(fila, 4).value = categoria.analisis or ""
        ws.cell(fila, 5).value = categoria.fecha_registro.strftime("%d/%m/%Y %H:%M")
        fila += 1

    # Ajustar ancho de columnas
    for columna in ws.columns:
        ancho = max(len(str(celda.value or "")) for celda in columna)
        ws.column_dimensions[columna[0].column_letter].width = ancho + 5

    # Activar filtros
    ws.auto_filter.ref = f"A1:E{fila-1}"

    return wb